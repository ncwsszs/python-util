import os
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook

# 起始日期
start_date = datetime(2024, 10, 1)
weeks = 44
template_path = "模板.xlsx"
output_dir = "输出文件"
os.makedirs(output_dir, exist_ok=True)

def rand_percent(low, high):
    return f"{round(random.uniform(low, high), 2):.2f}%"

def rand_g(low, high):
    return f"{round(random.uniform(low, high), 2):.2f}G"

for week in range(1, weeks + 1):
    # 日期范围
    week_start = start_date + timedelta(weeks=week - 1)
    week_end = week_start + timedelta(days=6)
    b3_value = f"{week_start.strftime('%Y.%m.%d')}-{week_end.strftime('%Y.%m.%d')}"

    # 生成随机值
    d15_value = rand_percent(2.0, 6.0)     # 百分比
    d16_value = rand_g(1.5, 1.9)           # G 值
    d17_value = rand_percent(28.0, 35.0)   # 百分比

    # 打开模板并写入
    wb = load_workbook(template_path)
    ws = wb.active

    ws["B3"].value = b3_value
    ws["D15"].value = d15_value
    ws["D16"].value = d16_value
    ws["D17"].value = d17_value

    # 保存新文件
    filename = f"运行维护-第{week}周.xlsx"
    wb.save(os.path.join(output_dir, filename))

print(f"✅ 共生成 {weeks} 份文件，保存在“{output_dir}/”目录下。")
