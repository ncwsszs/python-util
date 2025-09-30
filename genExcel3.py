import pandas as pd
from openpyxl import load_workbook
import os

# 读取Excel数据
df = pd.read_excel("问题处理数据.xlsx")

# 创建输出文件夹
output_dir = "问题处理与跟踪记录"
os.makedirs(output_dir, exist_ok=True)

# 跟踪日期文件名重复情况
filename_counter = {}

# 遍历每一行数据
for _, row in df.iterrows():
    时间 = row["时间"]
    事项 = row["事项"]
    描述 = row["描述"]
    提出人 = row["提出人"] if pd.notna(row["提出人"]) else ""
    工作量 = row["工作量(人天)"]
    开发 = row["开发"]
    处理过程 = row["处理过程"]
    处理结果 = row["处理结果"]

    # 加载模板
    wb = load_workbook("问题处理与跟踪记录.xlsx")
    ws = wb.active

    # 统一时间格式
    if isinstance(时间, pd.Timestamp):
        e4_date = 时间.strftime("%Y-%m-%d")
        filename_date = 时间.strftime("%Y年%m月%d日")
    else:
        e4_date = str(时间).split()[0]
        parts = e4_date.split("/")
        if len(parts) == 3:
            filename_date = f"{parts[0]}年{int(parts[1])}月{int(parts[2])}日"
        else:
            filename_date = e4_date

    # 写入数据到指定单元格
    ws["E3"] = str(开发)
    ws["E4"] = e4_date
    ws["C4"] = 工作量
    ws["C5"] = 描述
    ws["C6"] = 处理过程
    ws["C7"] = 处理结果
    ws["G3"] = 提出人

    # 构造基础文件名
    base_filename = f"问题处理与跟踪记录_{filename_date}"

    # 处理重复日期的文件名
    if base_filename not in filename_counter:
        filename_counter[base_filename] = 0
        filename = f"{base_filename}.xlsx"
    else:
        filename_counter[base_filename] += 1
        suffix = filename_counter[base_filename]
        filename = f"{base_filename}_{suffix}.xlsx"

    # 保存路径
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

print("✅ 所有Excel文件已生成，重复日期文件名已添加编号后缀。")
